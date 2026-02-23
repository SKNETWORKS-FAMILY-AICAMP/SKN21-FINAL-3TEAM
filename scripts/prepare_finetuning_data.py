"""
파인튜닝 데이터 준비 스크립트

judgment_raw.xlsx (1,000건) + regulation_qa_raw.xlsx (500건) → JSONL 변환
- 프로덕션 judgment_agent 시스템 프롬프트 + 사용자 프롬프트 + JSON 응답 포맷
- SFTTrainer chat format (messages 배열)
- 90/10 층화추출(stratified) train/eval 분할

사용법:
    python scripts/prepare_finetuning_data.py
"""

import json
import random
import re
import sys
from collections import Counter
from pathlib import Path

import openpyxl

# 프로젝트 루트
BASE_DIR = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(BASE_DIR))

from ai.llm.prompts import JUDGMENT_SYSTEM_PROMPT
from scripts.benchmark.regulation_texts import REGULATION_TEXTS, find_article_text

# ── 경로 ──
JUDGMENT_XLSX = BASE_DIR / "data" / "training" / "v1_judgment" / "judgment_raw.xlsx"
QA_XLSX = BASE_DIR / "data" / "training" / "v1_judgment" / "regulation_qa_raw.xlsx"
OUTPUT_DIR = BASE_DIR / "data" / "training" / "v1_judgment"
TRAIN_PATH = OUTPUT_DIR / "train.jsonl"
EVAL_PATH = OUTPUT_DIR / "eval.jsonl"

# ── 판단유형 매핑 ──
JUDGMENT_TYPE_MAP = {
    "Yes": "yes",
    "yes": "yes",
    "YES": "yes",
    "No": "no",
    "no": "no",
    "NO": "no",
    "조건부": "conditional",
}

# ── Confidence 기준 ──
JUDGMENT_CONFIDENCE = {"yes": 0.92, "no": 0.92, "conditional": 0.75}
QA_CONFIDENCE = {"yes": 0.88, "no": 0.88, "conditional": 0.72}

SEED = 42


def _extract_article_name(article_raw: str) -> str:
    """조항 컬럼에서 정규화된 조항명 추출. '제17조(정보시스템 접근통제)' → '제17조 (정보시스템 접근통제)'"""
    if not article_raw:
        return ""
    # 괄호 앞에 공백이 없으면 추가
    normalized = re.sub(r"(제\d+조)\(", r"\1 (", article_raw.strip())
    return normalized


def _build_user_message(article_name: str, regulation_text: str, question: str) -> str:
    """학습용 사용자 프롬프트 구성 (프로덕션 _build_user_prompt 간소화 버전)"""
    parts = []
    parts.append("## 관련 규정 문서")
    if article_name and regulation_text:
        parts.append(f"### {article_name}")
        parts.append(regulation_text)
    else:
        parts.append("관련 규정 문서를 찾지 못했습니다.")
    parts.append("")
    parts.append(f"## 사용자 질문")
    parts.append(question)
    return "\n".join(parts)


def _build_assistant_response(
    result: str,
    confidence: float,
    reasoning: str,
    article_name: str,
    regulation_content: str,
    conditions: str | None,
    alternatives: list[str],
) -> str:
    """학습용 assistant JSON 응답 구성"""
    regulations = []
    if article_name:
        regulations.append({
            "article": article_name,
            "relevance": "높음",
            "content": regulation_content or reasoning,
        })

    response = {
        "result": result,
        "confidence": confidence,
        "reasoning": reasoning,
        "regulations": regulations,
        "cross_references": [],
        "conditions": conditions,
        "alternatives": alternatives,
    }
    return json.dumps(response, ensure_ascii=False)


# ── Judgment 변환 ──


def convert_judgment(xlsx_path: Path) -> list[dict]:
    """judgment_raw.xlsx → messages 리스트 변환

    컬럼 매핑:
      col 1: ID
      col 2: 조항
      col 3: 판단유형 (Yes/No/조건부)
      col 4: 직급
      col 5: 부서
      col 6: 상황
      col 7: 질문(input)
      col 8: 근거
      col 9: 대안
      col 10: 전체출력(output)
    """
    wb = openpyxl.load_workbook(str(xlsx_path), read_only=True)
    ws = wb.active
    samples = []
    skipped = 0

    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not row or len(row) < 8:
            skipped += 1
            continue

        row_id = str(row[0] or "").strip()
        article_raw = str(row[1] or "").strip()
        judgment_type = str(row[2] or "").strip()
        question = str(row[6] or "").strip()
        reasoning_raw = str(row[7] or "").strip()
        alternative_raw = str(row[8] or "").strip() if len(row) > 8 and row[8] else ""

        # 필수 필드 검증
        if not question or not judgment_type:
            skipped += 1
            continue

        # 판단유형 매핑
        result = JUDGMENT_TYPE_MAP.get(judgment_type)
        if not result:
            print(f"  [WARN] row {row_idx}: 알 수 없는 판단유형 '{judgment_type}', 건너뜀")
            skipped += 1
            continue

        # 조항 원문 조회
        article_name = _extract_article_name(article_raw)
        regulation_text = find_article_text(article_raw)

        # confidence
        confidence = JUDGMENT_CONFIDENCE[result]

        # conditions / alternatives 분기
        conditions = None
        alternatives = []
        if result == "conditional" and alternative_raw:
            conditions = alternative_raw
        elif result == "no" and alternative_raw:
            alternatives = [alternative_raw]

        # 규정 content = 근거 텍스트 (reasoning과 다를 수 있음)
        regulation_content = reasoning_raw

        # user message 구성
        user_msg = _build_user_message(article_name, regulation_text, question)

        # assistant response 구성
        assistant_msg = _build_assistant_response(
            result=result,
            confidence=confidence,
            reasoning=reasoning_raw,
            article_name=article_name,
            regulation_content=regulation_content,
            conditions=conditions,
            alternatives=alternatives,
        )

        samples.append({
            "messages": [
                {"role": "system", "content": JUDGMENT_SYSTEM_PROMPT},
                {"role": "user", "content": user_msg},
                {"role": "assistant", "content": assistant_msg},
            ],
            "_result": result,  # 층화추출용 메타 (JSONL에는 미포함)
            "_source": "judgment",
            "_id": row_id,
        })

    wb.close()
    print(f"  Judgment: {len(samples)}건 변환 완료 (건너뜀: {skipped}건)")
    return samples


# ── QA 변환 ──

# 키워드 기반 판단 카테고리 자동 분류
_NO_KEYWORDS = ["불가", "금지", "안 됩니다", "할 수 없", "허용되지 않", "위반", "아니 됩니다"]
_YES_KEYWORDS = ["가능", "허용", "됩니다", "할 수 있", "맞습니다", "적용됩니다"]
_CONDITIONAL_KEYWORDS = ["조건", "단,", "다만", "경우에", "경우에만", "승인을 받", "사전에"]


def _classify_qa_result(answer: str) -> str:
    """답변 텍스트에서 키워드 기반으로 judgment 카테고리 자동 분류"""
    text = answer.strip()

    # no 키워드 먼저 체크 (더 명확한 판단)
    no_count = sum(1 for kw in _NO_KEYWORDS if kw in text)
    yes_count = sum(1 for kw in _YES_KEYWORDS if kw in text)
    cond_count = sum(1 for kw in _CONDITIONAL_KEYWORDS if kw in text)

    # conditional 키워드가 다수 + yes 키워드도 있으면 conditional
    if cond_count >= 2 and yes_count > 0:
        return "conditional"
    # no 키워드 우세
    if no_count > yes_count and no_count > 0:
        return "no"
    # conditional 키워드가 있으면
    if cond_count > 0 and no_count == 0:
        return "conditional"
    # yes 키워드가 있거나, 순수 설명형
    if yes_count > 0:
        return "yes"
    # 기본: 규정 존재 확인 → yes
    return "yes"


def convert_qa(xlsx_path: Path) -> list[dict]:
    """regulation_qa_raw.xlsx → messages 리스트 변환

    컬럼 매핑:
      col 1: No
      col 2: 조항
      col 3: 장(Chapter)
      col 4: 질문유형
      col 5: 직급
      col 6: 부서
      col 7: 상황
      col 8: Instruction
      col 9: Input (질문)
      col 10: Output (답변)
    """
    wb = openpyxl.load_workbook(str(xlsx_path), read_only=True)
    ws = wb.active
    samples = []
    skipped = 0

    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not row or len(row) < 10:
            skipped += 1
            continue

        row_no = str(row[0] or "").strip()
        article_raw = str(row[1] or "").strip()
        question = str(row[8] or "").strip()
        answer = str(row[9] or "").strip()

        # 필수 필드 검증
        if not question or not answer:
            skipped += 1
            continue

        # 조항 원문 조회
        article_name = _extract_article_name(article_raw)
        regulation_text = find_article_text(article_raw)

        # 키워드 기반 분류
        result = _classify_qa_result(answer)
        confidence = QA_CONFIDENCE.get(result, 0.88)

        # QA 답변 → reasoning으로 변환
        reasoning = answer

        # conditions / alternatives
        conditions = None
        alternatives = []
        if result == "conditional":
            # 답변에서 조건 부분 추출 시도
            conditions = answer
        elif result == "no":
            alternatives = []

        # user message
        user_msg = _build_user_message(article_name, regulation_text, question)

        # assistant response
        assistant_msg = _build_assistant_response(
            result=result,
            confidence=confidence,
            reasoning=reasoning,
            article_name=article_name,
            regulation_content=answer[:200],  # 답변 앞부분을 규정 content로
            conditions=conditions,
            alternatives=alternatives,
        )

        samples.append({
            "messages": [
                {"role": "system", "content": JUDGMENT_SYSTEM_PROMPT},
                {"role": "user", "content": user_msg},
                {"role": "assistant", "content": assistant_msg},
            ],
            "_result": result,
            "_source": "qa",
            "_id": f"QA-{row_no}",
        })

    wb.close()
    print(f"  QA: {len(samples)}건 변환 완료 (건너뜀: {skipped}건)")
    return samples


# ── Train/Eval 분할 (층화추출) ──


def stratified_split(
    samples: list[dict], eval_ratio: float = 0.1, seed: int = SEED
) -> tuple[list[dict], list[dict]]:
    """result 타입별 층화추출로 train/eval 분할"""
    random.seed(seed)

    # result별 그룹핑
    groups: dict[str, list[dict]] = {}
    for s in samples:
        r = s["_result"]
        groups.setdefault(r, []).append(s)

    train_set = []
    eval_set = []

    for result_type, group in groups.items():
        random.shuffle(group)
        n_eval = max(1, int(len(group) * eval_ratio))
        eval_set.extend(group[:n_eval])
        train_set.extend(group[n_eval:])

    # 최종 셔플
    random.shuffle(train_set)
    random.shuffle(eval_set)

    return train_set, eval_set


# ── JSONL 저장 ──


def save_jsonl(samples: list[dict], path: Path):
    """messages만 추출하여 JSONL 저장 (메타 필드 _result, _source, _id 제거)"""
    path.parent.mkdir(parents=True, exist_ok=True)
    with open(path, "w", encoding="utf-8") as f:
        for s in samples:
            record = {"messages": s["messages"]}
            f.write(json.dumps(record, ensure_ascii=False) + "\n")
    print(f"  저장: {path} ({len(samples)}건)")


# ── 검증 ──


def validate_samples(samples: list[dict], label: str):
    """전체 샘플 JSON 파싱 검증 + 통계 출력"""
    total = len(samples)
    json_ok = 0
    json_fail = 0
    has_regulation = 0
    result_counter = Counter()
    source_counter = Counter()

    for s in samples:
        msgs = s["messages"]
        # assistant 메시지 JSON 파싱 검증
        try:
            parsed = json.loads(msgs[2]["content"])
            json_ok += 1
            result_counter[parsed.get("result", "?")] += 1
        except (json.JSONDecodeError, IndexError):
            json_fail += 1

        # user 메시지에 규정 원문 포함 여부
        user_content = msgs[1]["content"]
        if "## 관련 규정 문서" in user_content and "관련 규정 문서를 찾지 못했습니다" not in user_content:
            has_regulation += 1

        source_counter[s.get("_source", "?")] += 1

    print(f"\n{'='*60}")
    print(f"  [{label}] 검증 결과")
    print(f"{'='*60}")
    print(f"  총 건수: {total}")
    print(f"  JSON 파싱 성공: {json_ok} / {total} ({json_ok/total*100:.1f}%)")
    if json_fail > 0:
        print(f"  JSON 파싱 실패: {json_fail}")
    print(f"  규정 원문 포함: {has_regulation} / {total} ({has_regulation/total*100:.1f}%)")
    print(f"  Result 분포: {dict(result_counter)}")
    print(f"  Source 분포: {dict(source_counter)}")
    print(f"{'='*60}")

    return json_fail == 0


def print_random_samples(samples: list[dict], n: int = 10):
    """랜덤 N건 샘플 출력으로 품질 확인"""
    random.seed(SEED + 1)
    picks = random.sample(samples, min(n, len(samples)))

    print(f"\n{'='*60}")
    print(f"  랜덤 {len(picks)}건 샘플 미리보기")
    print(f"{'='*60}")

    for i, s in enumerate(picks, 1):
        msgs = s["messages"]
        user_msg = msgs[1]["content"]
        assistant_msg = msgs[2]["content"]

        # 질문 부분만 추출
        q_match = re.search(r"## 사용자 질문\n(.+)", user_msg, re.DOTALL)
        question = q_match.group(1).strip()[:80] if q_match else user_msg[:80]

        # JSON 파싱
        try:
            parsed = json.loads(assistant_msg)
            result = parsed["result"]
            conf = parsed["confidence"]
            reasoning = parsed["reasoning"][:60]
        except Exception:
            result = "?"
            conf = 0
            reasoning = assistant_msg[:60]

        print(f"\n  [{i}] {s.get('_id', '?')} ({s.get('_source', '?')})")
        print(f"      질문: {question}")
        print(f"      판단: {result} (confidence={conf})")
        print(f"      근거: {reasoning}...")


# ── 메인 ──


def main():
    print("=" * 60)
    print("  파인튜닝 데이터 준비 시작")
    print("=" * 60)

    # 1. Judgment 변환
    print("\n[1/5] Judgment 데이터 변환...")
    judgment_samples = convert_judgment(JUDGMENT_XLSX)

    # 2. QA 변환
    print("\n[2/5] QA 데이터 변환...")
    qa_samples = convert_qa(QA_XLSX)

    # 3. 합치기 + 분할
    print("\n[3/5] Train/Eval 분할 (90/10, 층화추출)...")
    all_samples = judgment_samples + qa_samples
    print(f"  전체: {len(all_samples)}건")

    train_set, eval_set = stratified_split(all_samples, eval_ratio=0.1)
    print(f"  Train: {len(train_set)}건, Eval: {len(eval_set)}건")

    # 4. JSONL 저장
    print("\n[4/5] JSONL 저장...")
    save_jsonl(train_set, TRAIN_PATH)
    save_jsonl(eval_set, EVAL_PATH)

    # 5. 검증
    print("\n[5/5] 검증...")
    all_ok = True
    all_ok &= validate_samples(train_set, "Train")
    all_ok &= validate_samples(eval_set, "Eval")
    validate_samples(all_samples, "전체")

    # 랜덤 샘플 출력
    print_random_samples(all_samples, n=10)

    if all_ok:
        print(f"\n  모든 검증 통과!")
    else:
        print(f"\n  일부 검증 실패 — 위 로그 확인")

    print(f"\n  출력 파일:")
    print(f"    - {TRAIN_PATH}")
    print(f"    - {EVAL_PATH}")


if __name__ == "__main__":
    main()
