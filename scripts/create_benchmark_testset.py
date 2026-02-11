"""
모델 벤치마크용 테스트 데이터셋 생성 스크립트
- judgment_raw.xlsx에서 21개 샘플링 (Yes 7 / No 7 / 조건부 7)
- regulation_qa_raw.xlsx에서 16개 샘플링 (질문유형별 2개씩)
- proceedings/*.json에서 16개 샘플링 (파일별 2개씩)
- 한국어 이해력 + 구조화 출력 테스트 17개 직접 작성
=> 총 70개
"""

import json
import random
import os
import sys
import openpyxl

sys.stdout.reconfigure(encoding="utf-8")
random.seed(42)

BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
OUTPUT_PATH = os.path.join(BASE_DIR, "data", "evaluation", "benchmark_testset.jsonl")

testset = []
test_id = 1


# ============================================================
# 1. 규정 판단 (Judgment) - 21개
# ============================================================
print("[1/4] 규정 판단 샘플링...")

wb = openpyxl.load_workbook(
    os.path.join(BASE_DIR, "data", "training", "v1_judgment", "judgment_raw.xlsx"),
    read_only=True,
)
ws = wb["판단예시_1000"]

judgment_by_type = {"Yes": [], "No": [], "조건부": []}
for i, row in enumerate(ws.iter_rows(values_only=True)):
    if i == 0:
        continue
    jtype = row[2]
    if jtype in judgment_by_type:
        judgment_by_type[jtype].append(
            {
                "id": row[0],
                "article": row[1],
                "type": row[2],
                "rank": row[3],
                "dept": row[4],
                "situation": row[5],
                "question": row[6],
                "basis": row[7],
                "alternative": row[8],
                "full_output": row[9],
            }
        )

for jtype, items in judgment_by_type.items():
    # 다양한 조항에서 뽑기 위해 조항별 그룹핑 후 하나씩 선택
    by_article = {}
    for item in items:
        by_article.setdefault(item["article"], []).append(item)

    articles = list(by_article.keys())
    random.shuffle(articles)
    sampled = []
    for article in articles:
        if len(sampled) >= 7:
            break
        sampled.append(random.choice(by_article[article]))

    for item in sampled:
        # 실서비스와 동일하게 규정 원문을 input에 포함 (RAG가 붙여주는 형태)
        regulation_context = f"관련 규정:\n{item['article']}"
        if item["basis"]:
            regulation_context += f": {item['basis']}"
        bench_input = f"{regulation_context}\n\n질문: {item['question']}"

        testset.append(
            {
                "test_id": f"BENCH-{test_id:03d}",
                "category": "judgment",
                "subcategory": item["type"],
                "instruction": "당신은 듀듀테크놀로지 사내규정 전문가입니다. 다음 규정과 질문을 읽고 Yes/No/조건부로 판단하고, 근거와 대안을 제시하세요.",
                "input": bench_input,
                "reference_output": item["full_output"],
                "metadata": {
                    "source_id": item["id"],
                    "article": item["article"],
                    "judgment_type": item["type"],
                },
                "eval_criteria": [
                    "판단 정확도 (Yes/No/조건부)",
                    "근거 조항 언급 여부",
                    "응답 형식 준수",
                ],
            }
        )
        test_id += 1

print(f"  => {len([t for t in testset if t['category'] == 'judgment'])}개 생성")


# ============================================================
# 2. 규정 해석 Q&A - 16개
# ============================================================
print("[2/4] 규정 해석 Q&A 샘플링...")

wb2 = openpyxl.load_workbook(
    os.path.join(
        BASE_DIR, "data", "training", "v1_judgment", "regulation_qa_raw.xlsx"
    ),
    read_only=True,
)
ws2 = wb2["규정해석_QA_500"]

qa_by_type = {}
for i, row in enumerate(ws2.iter_rows(values_only=True)):
    if i == 0:
        continue
    qtype = row[3]  # 질문유형
    if qtype not in qa_by_type:
        qa_by_type[qtype] = []
    qa_by_type[qtype].append(
        {
            "no": row[0],
            "article": row[1],
            "chapter": row[2],
            "qtype": row[3],
            "rank": row[4],
            "dept": row[5],
            "situation": row[6],
            "instruction": row[7],
            "input_q": row[8],
            "output_a": row[9],
        }
    )

# 주요 질문유형에서 2개씩 (8개 유형 x 2 = 16개)
target_types = [
    "상황설정",
    "직접",
    "우회",
    "절차확인",
    "위반사례",
    "예외케이스",
    "복합조항",
    "비교",
]
for qtype in target_types:
    if qtype not in qa_by_type:
        continue
    items = qa_by_type[qtype]
    sampled = random.sample(items, min(2, len(items)))
    for item in sampled:
        # 실서비스와 동일하게 규정 조항을 input에 포함 (RAG가 붙여주는 형태)
        qa_input = f"관련 규정: {item['article']}\n\n질문: {item['input_q']}"

        testset.append(
            {
                "test_id": f"BENCH-{test_id:03d}",
                "category": "regulation_qa",
                "subcategory": item["qtype"],
                "instruction": item["instruction"],
                "input": qa_input,
                "reference_output": item["output_a"],
                "metadata": {
                    "source_no": item["no"],
                    "article": item["article"],
                    "chapter": item["chapter"],
                    "question_type": item["qtype"],
                },
                "eval_criteria": [
                    "답변 정확성 (규정 내용과 일치)",
                    "설명 충실도 (예시 포함 여부)",
                    "한국어 자연스러움",
                ],
            }
        )
        test_id += 1

print(
    f"  => {len([t for t in testset if t['category'] == 'regulation_qa'])}개 생성"
)


# ============================================================
# 3. 회의록 분석 (Meeting Analysis) - 16개
# ============================================================
print("[3/4] 회의록 분석 샘플링...")

proc_dir = os.path.join(BASE_DIR, "data", "proceedings")
proc_files = sorted([f for f in os.listdir(proc_dir) if f.endswith(".json")])

for fname in proc_files:
    with open(os.path.join(proc_dir, fname), "r", encoding="utf-8") as f:
        data = json.load(f)

    sampled = random.sample(data, min(2, len(data)))
    for item in sampled:
        # output이 문자열인 경우 JSON 파싱
        output = item.get("output", "")
        if isinstance(output, str):
            try:
                output = json.loads(output)
            except json.JSONDecodeError:
                pass

        testset.append(
            {
                "test_id": f"BENCH-{test_id:03d}",
                "category": "meeting_analysis",
                "subcategory": fname.replace(".json", ""),
                "instruction": item.get(
                    "instruction",
                    "다음 회의록에서 회의 정보, 결정사항, Action Item, 참석자를 추출하세요.",
                ),
                "input": item.get("input", ""),
                "reference_output": output
                if isinstance(output, str)
                else json.dumps(output, ensure_ascii=False),
                "metadata": {"source_file": fname},
                "eval_criteria": [
                    "JSON 구조화 능력",
                    "핵심 정보 추출 정확도 (회의정보, 결정사항, action_items, 참석자)",
                    "한국어 요약 품질",
                ],
            }
        )
        test_id += 1

print(
    f"  => {len([t for t in testset if t['category'] == 'meeting_analysis'])}개 생성"
)


# ============================================================
# 4. 한국어 이해력 + 구조화 출력 테스트 - 17개
# ============================================================
print("[4/4] 한국어 이해력 + 구조화 출력 테스트 추가...")

extra_tests = [
    # --- 문서 요약 (5개) ---
    {
        "category": "doc_summary",
        "subcategory": "보고서형",
        "instruction": "주어진 필드에 맞게 문서를 요약하세요.",
        "input": "필드: title, summary, key_points, conclusion\n\n원본 문서:\n제목: 2026년 1분기 IT 인프라 점검 보고서\n\n1. 점검 개요\n점검 기간: 2026.01.13~01.24\n대상: 전사 서버 48대, 네트워크 장비 120대\n\n2. 주요 발견사항\n- 서버 가동률 99.7% (목표 99.5% 초과 달성)\n- 디스크 사용률 80% 초과 서버 5대 발견\n- 방화벽 정책 미갱신 3건\n- SSL 인증서 만료 예정 2건 (3월 내)\n\n3. 조치 계획\n디스크 증설 및 방화벽 정책 갱신 예정",
        "reference_output": '{"title": "2026년 1분기 IT 인프라 점검 보고서", "summary": "서버 가동률 99.7%로 목표 달성. 디스크 사용률 초과 서버 5대, 방화벽 미갱신 3건 등 일부 조치 필요", "key_points": "- 서버 48대, 네트워크 장비 120대 점검\\n- 디스크 80% 초과 서버 5대\\n- 방화벽 정책 미갱신 3건\\n- SSL 인증서 만료 예정 2건", "conclusion": "전반적 안정성 양호. 디스크 증설 및 방화벽 정책 갱신, SSL 인증서 갱신 조치 필요"}',
        "eval_criteria": [
            "JSON 출력 형식 준수",
            "필드별 요약 정확도",
            "핵심 수치 포함 여부",
        ],
    },
    {
        "category": "doc_summary",
        "subcategory": "기획서형",
        "instruction": "주어진 필드에 맞게 문서를 요약하세요.",
        "input": "필드: title, background, summary, key_points\n\n원본 문서:\n제목: 사내 AI 챗봇 도입 기획서\n\n1. 배경\n현재 인사/IT/총무 부서에 월 평균 620건의 반복 문의가 발생하고 있으며, 담당자 1인당 일 평균 45분을 문의 응대에 사용하고 있습니다.\n\n2. 제안\nLLM 기반 사내 챗봇을 도입하여 규정 질문, 일정 관리, 문서 작성을 자동화합니다.\n주요 기능: 규정 판단, 문서 요약/생성, 일정 등록, 회의록 정리\n\n3. 기대 효과\n- 반복 문의 70% 자동 처리\n- 담당자 업무시간 주당 3.5시간 절감\n- 신입사원 온보딩 기간 단축",
        "reference_output": '{"title": "사내 AI 챗봇 도입 기획서", "background": "인사/IT/총무 부서에 월 620건 반복 문의 발생, 담당자 1인당 일 45분 응대에 소요", "summary": "LLM 기반 사내 챗봇으로 규정 질문, 일정 관리, 문서 작성 자동화를 제안하는 기획", "key_points": "- 규정 판단, 문서 요약/생성, 일정 등록, 회의록 정리 기능\\n- 반복 문의 70% 자동 처리 목표\\n- 주당 3.5시간 업무시간 절감 기대"}',
        "eval_criteria": [
            "JSON 출력 형식 준수",
            "배경 정보 정확 요약",
            "핵심 수치 포함 여부",
        ],
    },
    {
        "category": "doc_summary",
        "subcategory": "감사형",
        "instruction": "주어진 필드에 맞게 문서를 요약하세요.",
        "input": "필드: title, summary, key_points, action_required\n\n원본 문서:\n제목: 2025년 하반기 개인정보 처리 감사 보고서\n\n1. 감사 범위: 전사 개인정보 처리 현황 (고객DB, 직원DB, 협력사DB)\n2. 주요 결과\n- 고객 DB 접근 로그 미보관: 마케팅팀 3건\n- 퇴사자 개인정보 미파기: 인사팀 8건 (30일 초과 보관)\n- 동의 없는 제3자 제공: 영업팀 1건\n- 암호화 미적용: 협력사 DB 연동 구간 1건\n3. 전체 준수율: 87% (전년 대비 +5%p)",
        "reference_output": '{"title": "2025년 하반기 개인정보 처리 감사 보고서", "summary": "전사 개인정보 처리 감사 결과 준수율 87%. 접근로그 미보관, 퇴사자 정보 미파기 등 13건 미준수 발견", "key_points": "- 고객 DB 접근 로그 미보관 3건 (마케팅팀)\\n- 퇴사자 개인정보 미파기 8건 (인사팀)\\n- 동의 없는 제3자 제공 1건 (영업팀)\\n- 암호화 미적용 1건 (협력사 DB)", "action_required": "- 마케팅팀: 접근 로그 보관 체계 구축\\n- 인사팀: 퇴사자 정보 파기 프로세스 점검\\n- 영업팀: 제3자 제공 동의 절차 재정비\\n- IT팀: 협력사 DB 구간 암호화 적용"}',
        "eval_criteria": [
            "JSON 출력 형식 준수",
            "부서별 이슈 정확 분류",
            "조치사항 구체성",
        ],
    },
    {
        "category": "doc_summary",
        "subcategory": "공지형",
        "instruction": "주어진 필드에 맞게 문서를 요약하세요.",
        "input": "필드: title, summary, recipients, action_required\n\n원본 문서:\n[공지] 2026년 상반기 정보보호 교육 안내\n\n전 직원 대상으로 정보보호 교육을 실시합니다.\n\n1. 교육 기간: 2026.03.03 ~ 03.14 (2주간)\n2. 대상: 전 직원 (정규직, 계약직, 인턴 포함)\n3. 방법: 온라인 (사내 LMS)\n4. 소요시간: 약 2시간\n5. 필수 이수: 미이수 시 인사 불이익\n6. 문의: 보안팀 (내선 1234)",
        "reference_output": '{"title": "2026년 상반기 정보보호 교육 안내", "summary": "2026.03.03~03.14 전 직원 대상 온라인 정보보호 교육 실시. 약 2시간 소요, 미이수 시 인사 불이익", "recipients": "전 직원 (정규직, 계약직, 인턴 포함)", "action_required": "- 기간 내 LMS에서 정보보호 교육 이수 (2시간)\\n- 미이수 시 인사 불이익 있으므로 반드시 완료"}',
        "eval_criteria": [
            "JSON 출력 형식 준수",
            "일정/대상 정확도",
            "필수 조치사항 포함",
        ],
    },
    {
        "category": "doc_summary",
        "subcategory": "기술문서형",
        "instruction": "주어진 필드에 맞게 문서를 요약하세요.",
        "input": "필드: title, summary, key_points, technical_details\n\n원본 문서:\n제목: API 게이트웨이 마이그레이션 계획서\n\n1. 현황: 레거시 API 게이트웨이(Kong 2.x)의 EOL이 2026년 6월로 예정\n2. 목표: Kong 3.x + Kubernetes Ingress로 전환\n3. 영향 범위: 내부 API 47개, 외부 연동 API 12개, 일 평균 트래픽 150만 req\n4. 마이그레이션 전략\n- Blue-Green 배포 방식\n- 단계별 전환 (내부 API → 외부 API 순서)\n- 롤백 계획 포함\n5. 일정: 4월 PoC → 5월 내부 전환 → 6월 외부 전환",
        "reference_output": '{"title": "API 게이트웨이 마이그레이션 계획서", "summary": "Kong 2.x EOL 대응으로 Kong 3.x + K8s Ingress 전환 계획. Blue-Green 배포로 단계별 마이그레이션", "key_points": "- 레거시 Kong 2.x → Kong 3.x + Kubernetes Ingress\\n- 내부 47개 + 외부 12개 API 대상\\n- 4월 PoC → 5월 내부 → 6월 외부 전환", "technical_details": "- Blue-Green 배포 방식 적용\\n- 일 평균 트래픽 150만 req 처리 필요\\n- 내부 API 우선 전환 후 외부 API 순차 적용\\n- 롤백 계획 포함"}',
        "eval_criteria": [
            "JSON 출력 형식 준수",
            "기술 용어 정확성",
            "일정/수치 포함 여부",
        ],
    },
    # --- 리스크 감지 (5개) ---
    {
        "category": "risk_detection",
        "subcategory": "위반_접근권한",
        "instruction": "다음 내용에서 규정 위반 여부를 감지하세요.",
        "input": "회의 내용: 인턴 박지호에게 프로덕션 DB 직접 접근 권한을 부여하기로 결정함. 개발 속도를 위해 필요하다는 팀장 의견 반영.\n\n관련 규정:\n제18조(데이터베이스 접근통제) 제3항 - 수습기간 중인 직원은 프로덕션 데이터베이스에 직접 접근할 수 없다.",
        "reference_output": '{"risk_detected": true, "risk_level": "high", "violation": "제18조(데이터베이스 접근통제) 제3항 위반", "detail": "수습기간 중인 인턴에게 프로덕션 DB 직접 접근 권한 부여 시도", "recommendation": "테스트 DB 환경을 제공하고, 수습 종료 후 접근 권한 재검토"}',
        "eval_criteria": [
            "위반 감지 정확도",
            "위반 조항 정확 인용",
            "대안 제시 적절성",
        ],
    },
    {
        "category": "risk_detection",
        "subcategory": "위반_비밀유지",
        "instruction": "다음 내용에서 규정 위반 여부를 감지하세요.",
        "input": "회의 내용: 마케팅팀 김대리가 경쟁사 이직 예정자와 진행 중인 프로젝트의 기술 스펙을 공유한 사실이 확인됨.\n\n관련 규정:\n제11조(비밀유지의무) 제1항 - 재직 중 취득한 회사의 영업비밀 및 기술정보를 외부에 유출하여서는 안 된다.",
        "reference_output": '{"risk_detected": true, "risk_level": "high", "violation": "제11조(비밀유지의무) 제1항 위반", "detail": "진행 중인 프로젝트 기술 스펙을 경쟁사 이직 예정자에게 유출", "recommendation": "즉시 해당 직원 면담 및 유출 범위 파악, 보안팀 보고 후 징계 절차 검토"}',
        "eval_criteria": [
            "위반 감지 정확도",
            "심각도 판단 적절성",
            "후속 조치 구체성",
        ],
    },
    {
        "category": "risk_detection",
        "subcategory": "정상_접근권한",
        "instruction": "다음 내용에서 규정 위반 여부를 감지하세요.",
        "input": "회의 내용: 정규직 백엔드 개발자 이수진에게 담당 프로젝트의 스테이징 DB 읽기 권한을 부여하기로 함. 부서장 승인 완료.\n\n관련 규정:\n제17조(정보시스템 접근통제) 제1항 - 업무 수행에 필요한 최소한의 접근 권한을 부여한다.\n제17조(정보시스템 접근통제) 제3항 - 접근 권한 부여 시 부서장의 승인을 받아야 한다.",
        "reference_output": '{"risk_detected": false, "risk_level": "none", "violation": "", "detail": "정규직 직원에게 담당 프로젝트 범위 내 최소 권한(스테이징 DB 읽기) 부여, 부서장 승인 완료로 절차 준수", "recommendation": ""}',
        "eval_criteria": [
            "정상 판단 정확도",
            "규정 준수 근거 설명",
            "불필요한 위험 경고 없음",
        ],
    },
    {
        "category": "risk_detection",
        "subcategory": "위반_원격근무",
        "instruction": "다음 내용에서 규정 위반 여부를 감지하세요.",
        "input": "회의 내용: 재택근무 중인 직원들이 카페에서 공용 Wi-Fi로 사내 시스템에 접속하는 사례가 보고됨. VPN 없이 직접 접속한 경우도 있음.\n\n관련 규정:\n제9조(원격근무) 제2항 - 원격근무 시 반드시 회사 승인 VPN을 통해 접속해야 한다.\n제9조(원격근무) 제4항 - 공공장소에서의 업무 수행 시 화면 보안필터를 사용해야 한다.",
        "reference_output": '{"risk_detected": true, "risk_level": "medium", "violation": "제9조(원격근무) 제2항 위반", "detail": "공용 Wi-Fi에서 VPN 미사용으로 사내 시스템 직접 접속", "recommendation": "전 직원 대상 원격근무 보안 수칙 재교육, VPN 미접속 시 사내 시스템 차단 정책 적용 검토"}',
        "eval_criteria": [
            "위반 감지 정확도",
            "관련 조항 모두 식별",
            "재발 방지 제안 포함",
        ],
    },
    {
        "category": "risk_detection",
        "subcategory": "정상_교육",
        "instruction": "다음 내용에서 규정 위반 여부를 감지하세요.",
        "input": "회의 내용: 보안팀에서 전 직원 대상 분기별 정보보호 교육을 정상 실시함. 이수율 98%, 미이수 2명은 출장 중으로 다음 주 보충 교육 예정.\n\n관련 규정:\n제29조(정보보호 교육) 제1항 - 전 직원은 분기별 1회 이상 정보보호 교육을 이수해야 한다.\n제29조(정보보호 교육) 제3항 - 부득이한 사유로 미이수 시 2주 이내 보충 교육을 이수해야 한다.",
        "reference_output": '{"risk_detected": false, "risk_level": "none", "violation": "", "detail": "분기별 교육 정상 실시(이수율 98%), 미이수자 2명은 출장 사유로 2주 이내 보충 교육 예정이므로 규정 준수", "recommendation": ""}',
        "eval_criteria": [
            "정상 판단 정확도",
            "보충 교육 규정 이해",
            "과잉 경고 없음",
        ],
    },
    # --- 한국어 이해력 (7개) ---
    {
        "category": "korean_understanding",
        "subcategory": "문맥추론",
        "instruction": "다음 문장을 읽고 질문에 답하세요.",
        "input": "회사 규정에 따르면, 연차는 입사 1년 후부터 15일이 부여되며, 3년 이상 근속 시 2년마다 1일씩 가산됩니다. 김대리는 입사 7년차입니다.\n\n질문: 김대리의 연차 일수는 며칠인가요?",
        "reference_output": "김대리는 입사 7년차이므로 기본 15일에 가산 연차를 더합니다. 3년 이상 근속 시 2년마다 1일 가산이므로, 3~4년 +1일, 5~6년 +1일 = 총 2일 가산. 따라서 15 + 2 = 17일입니다.",
        "eval_criteria": ["계산 정확도", "논리적 추론 과정", "한국어 설명 자연스러움"],
    },
    {
        "category": "korean_understanding",
        "subcategory": "요약",
        "instruction": "다음 내용을 3줄 이내로 요약하세요.",
        "input": "듀듀테크놀로지는 2024년 설립된 IT 기업으로, 사내 업무 자동화 솔루션을 개발하고 있습니다. 주요 제품은 AI 기반 사내 챗봇으로, 규정 질문 답변, 문서 자동 생성, 일정 관리 등의 기능을 제공합니다. 현재 직원 수는 약 50명이며, 서울 강남구에 본사를 두고 있습니다. 최근 시리즈 A 투자를 유치하여 사업 확장을 준비 중입니다.",
        "reference_output": "듀듀테크놀로지는 2024년 설립된 IT 기업으로 AI 기반 사내 챗봇(규정 답변, 문서 생성, 일정 관리)을 개발합니다. 직원 50명 규모이며 서울 강남구에 위치합니다. 최근 시리즈 A 투자를 유치하여 사업 확장을 준비 중입니다.",
        "eval_criteria": ["핵심 정보 포함", "3줄 이내 준수", "불필요한 정보 제거"],
    },
    {
        "category": "korean_understanding",
        "subcategory": "지시따르기",
        "instruction": "아래 조건에 맞게 이메일을 작성하세요.",
        "input": "조건:\n- 수신: 개발팀 전체\n- 발신: 보안팀 박팀장\n- 내용: 다음 주 수요일(3/5) 오후 2시 보안 패치 적용으로 인해 사내 시스템 2시간 점검 예정\n- 톤: 공식적, 정중",
        "reference_output": "제목: [공지] 사내 시스템 정기 점검 안내 (3/5 수요일)\n\n개발팀 여러분께,\n\n안녕하세요. 보안팀 박팀장입니다.\n\n다음 주 수요일(3월 5일) 오후 2시부터 약 2시간 동안 보안 패치 적용을 위한 사내 시스템 정기 점검이 진행될 예정입니다.\n\n점검 시간 동안 사내 시스템 접속이 일시적으로 제한될 수 있으니, 중요 작업은 점검 전에 저장해 주시기 바랍니다.\n\n불편을 드려 죄송하며, 문의 사항이 있으시면 보안팀으로 연락 부탁드립니다.\n\n감사합니다.\n보안팀 박팀장 드림",
        "eval_criteria": [
            "조건 충족 (일시, 내용, 톤)",
            "이메일 형식 준수",
            "한국어 비즈니스 문체",
        ],
    },
    {
        "category": "korean_understanding",
        "subcategory": "분류",
        "instruction": "다음 문장들의 의도를 분류하세요. 가능한 의도: judgment, doc_search, doc_generate, meeting_generate, schedule_add, schedule_view, general",
        "input": '1. "인턴이 프로덕션 서버에 접속해도 되나요?"\n2. "이번 주 일정 알려줘"\n3. "오늘 회의 내용 정리해줘"\n4. "보안 규정 제17조 내용이 뭐야?"\n5. "이 내용으로 보고서 만들어줘"',
        "reference_output": "1. judgment - 규정상 허용 여부 판단 요청\n2. schedule_view - 일정 조회 요청\n3. meeting_generate - 회의록 정리 요청\n4. doc_search - 규정 내용 검색 요청\n5. doc_generate - 문서 생성 요청",
        "eval_criteria": [
            "5개 분류 정확도",
            "분류 근거 설명",
            "경계 케이스 처리",
        ],
    },
    {
        "category": "korean_understanding",
        "subcategory": "논리판단",
        "instruction": "다음 상황을 읽고 논리적으로 판단하세요.",
        "input": "A팀 팀장이 퇴사 예정인 B 직원의 업무 인수인계를 위해, B 직원의 모든 시스템 접근 권한을 C 직원(신입)에게 일괄 이전하려 합니다.\n\n질문: 이 조치에 어떤 문제가 있을 수 있나요?",
        "reference_output": "여러 가지 문제가 있습니다:\n\n1. 최소 권한 원칙 위반: B 직원의 모든 권한을 일괄 이전하면 C 직원(신입)에게 업무에 불필요한 권한까지 부여될 수 있습니다.\n2. 신입사원 권한 제한: 수습기간 중인 신입사원에게 고급 권한(프로덕션 DB 등)이 부여될 위험이 있습니다.\n3. 권한 검토 누락: 개별 권한을 검토하지 않고 일괄 이전하면 보안 취약점이 발생할 수 있습니다.\n\n권장 조치: 업무별로 필요한 권한만 선별 부여하고, 부서장 및 보안팀 승인을 받아야 합니다.",
        "eval_criteria": [
            "보안 관점 문제 식별",
            "다각적 분석",
            "실질적 대안 제시",
        ],
    },
    {
        "category": "korean_understanding",
        "subcategory": "존댓말변환",
        "instruction": "다음 반말 문장을 비즈니스 존댓말로 변환하세요.",
        "input": '1. "이거 내일까지 해야 돼"\n2. "그 파일 나한테 보내줘"\n3. "회의 시간 바꿔야 할 것 같아"\n4. "이번 달 실적 별로야"',
        "reference_output": '1. "해당 건은 내일까지 완료해야 합니다."\n2. "해당 파일을 저에게 전달해 주시겠습니까?"\n3. "회의 시간을 변경해야 할 것 같습니다."\n4. "이번 달 실적이 다소 저조한 상황입니다."',
        "eval_criteria": [
            "존댓말 변환 정확성",
            "비즈니스 문체 적절성",
            "원래 의미 유지",
        ],
    },
    {
        "category": "korean_understanding",
        "subcategory": "다중조건추론",
        "instruction": "다음 조건을 모두 고려하여 답하세요.",
        "input": "조건:\n- 팀장 이상은 VPN 없이 사내 시스템 접속 가능 (긴급 시)\n- 대리 이하는 반드시 VPN 필요\n- 해외 출장 시 모든 직급이 VPN + MFA 필수\n- 현재 상황: 과장 김철수가 해외 출장 중 호텔에서 VPN 없이 접속 시도\n\n질문: 김 과장의 접속이 허용되나요? 이유를 설명하세요.",
        "reference_output": "허용되지 않습니다.\n\n김 과장은 과장급이므로 일반적으로는 팀장 이상 조건에 해당하지 않아 VPN이 필요합니다. 또한 현재 해외 출장 중이므로, 직급과 관계없이 VPN + MFA가 모두 필수입니다.\n\n따라서 두 가지 조건 모두에서 VPN 없는 접속은 불허됩니다:\n1. 과장급은 팀장 이상이 아니므로 VPN 필수\n2. 해외 출장 시 모든 직급 VPN + MFA 필수",
        "eval_criteria": [
            "다중 조건 정확 적용",
            "논리적 추론 과정",
            "최종 판단 정확성",
        ],
    },
]

for item in extra_tests:
    testset.append(
        {
            "test_id": f"BENCH-{test_id:03d}",
            "category": item["category"],
            "subcategory": item["subcategory"],
            "instruction": item["instruction"],
            "input": item["input"],
            "reference_output": item["reference_output"],
            "metadata": {"source": "manual"},
            "eval_criteria": item["eval_criteria"],
        }
    )
    test_id += 1

extra_categories = {"doc_summary", "risk_detection", "korean_understanding"}
for cat in extra_categories:
    count = len([t for t in testset if t["category"] == cat])
    print(f"  => {cat}: {count}개 생성")


# ============================================================
# 저장
# ============================================================
os.makedirs(os.path.dirname(OUTPUT_PATH), exist_ok=True)
with open(OUTPUT_PATH, "w", encoding="utf-8") as f:
    for item in testset:
        f.write(json.dumps(item, ensure_ascii=False) + "\n")

print(f"\n{'='*50}")
print(f"총 {len(testset)}개 테스트 항목 생성 완료")
print(f"저장: {OUTPUT_PATH}")

# 카테고리별 요약
print(f"\n카테고리별 분포:")
cats = {}
for item in testset:
    cats[item["category"]] = cats.get(item["category"], 0) + 1
for k, v in sorted(cats.items()):
    print(f"  {k}: {v}개")
