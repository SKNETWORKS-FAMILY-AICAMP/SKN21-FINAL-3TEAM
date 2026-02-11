"""
모델 벤치마크용 테스트 데이터셋 생성 스크립트 (v2 — 2026-02-11)
- judgment_raw.xlsx에서 21개 샘플링 (Yes 7 / No 7 / 조건부 7)
- regulation_qa_raw.xlsx에서 16개 샘플링 (질문유형별 2개씩) + 규정 원문 포함
- proceedings/*.json에서 16개 샘플링 (파일별 2개씩)
- 한국어 이해력 + 구조화 출력 테스트 34개 직접 작성
  - doc_summary 12개 / risk_detection 12개 / korean_understanding 10개
=> 총 87개
"""

import json
import random
import os
import sys
import openpyxl

from regulation_texts import find_article_text

sys.stdout.reconfigure(encoding="utf-8")
random.seed(42)

BASE_DIR = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
OUTPUT_PATH = os.path.join(BASE_DIR, "data", "evaluation", "benchmark_testset.jsonl")

testset = []
test_id = 1


# ============================================================
# 1. 규정 판단 (Judgment) - 21개
# ============================================================
print("[1/4] 규정 판단 샘플링...")

wb = openpyxl.load_workbook(
    os.path.join(BASE_DIR, "data", "training", "v1_judgment", "judgment_raw.xlsx"),
    read_only=True,
)
ws = wb["판단예시_1000"]

judgment_by_type = {"Yes": [], "No": [], "조건부": []}
for i, row in enumerate(ws.iter_rows(values_only=True)):
    if i == 0:
        continue
    jtype = row[2]
    if jtype in judgment_by_type:
        judgment_by_type[jtype].append(
            {
                "id": row[0],
                "article": row[1],
                "type": row[2],
                "rank": row[3],
                "dept": row[4],
                "situation": row[5],
                "question": row[6],
                "basis": row[7],
                "alternative": row[8],
                "full_output": row[9],
            }
        )

for jtype, items in judgment_by_type.items():
    # 다양한 조항에서 뽑기 위해 조항별 그룹핑 후 하나씩 선택
    by_article = {}
    for item in items:
        by_article.setdefault(item["article"], []).append(item)

    articles = list(by_article.keys())
    random.shuffle(articles)
    sampled = []
    for article in articles:
        if len(sampled) >= 7:
            break
        sampled.append(random.choice(by_article[article]))

    for item in sampled:
        # 실서비스와 동일하게 규정 원문 전문을 input에 포함 (RAG가 붙여주는 형태)
        article_full = find_article_text(item["article"])
        regulation_context = f"관련 규정:\n{item['article']}"
        if article_full:
            regulation_context += f"\n{article_full}"
        elif item["basis"]:
            regulation_context += f": {item['basis']}"
        bench_input = f"{regulation_context}\n\n질문: {item['question']}"

        testset.append(
            {
                "test_id": f"BENCH-{test_id:03d}",
                "category": "judgment",
                "subcategory": item["type"],
                "instruction": "당신은 듀듀테크놀로지 사내규정 전문가입니다. 다음 규정과 질문을 읽고 Yes/No/조건부로 판단하고, 근거와 대안을 제시하세요.",
                "input": bench_input,
                "reference_output": item["full_output"],
                "metadata": {
                    "source_id": item["id"],
                    "article": item["article"],
                    "judgment_type": item["type"],
                },
                "eval_criteria": [
                    "판단 정확도 (Yes/No/조건부)",
                    "근거 조항 언급 여부",
                    "응답 형식 준수",
                ],
            }
        )
        test_id += 1

print(f"  => {len([t for t in testset if t['category'] == 'judgment'])}개 생성")


# ============================================================
# 2. 규정 해석 Q&A - 16개
# ============================================================
print("[2/4] 규정 해석 Q&A 샘플링...")

wb2 = openpyxl.load_workbook(
    os.path.join(
        BASE_DIR, "data", "training", "v1_judgment", "regulation_qa_raw.xlsx"
    ),
    read_only=True,
)
ws2 = wb2["규정해석_QA_500"]

qa_by_type = {}
for i, row in enumerate(ws2.iter_rows(values_only=True)):
    if i == 0:
        continue
    qtype = row[3]  # 질문유형
    if qtype not in qa_by_type:
        qa_by_type[qtype] = []
    qa_by_type[qtype].append(
        {
            "no": row[0],
            "article": row[1],
            "chapter": row[2],
            "qtype": row[3],
            "rank": row[4],
            "dept": row[5],
            "situation": row[6],
            "instruction": row[7],
            "input_q": row[8],
            "output_a": row[9],
        }
    )

# 주요 질문유형에서 2개씩 (8개 유형 x 2 = 16개)
target_types = [
    "상황설정",
    "직접",
    "우회",
    "절차확인",
    "위반사례",
    "예외케이스",
    "복합조항",
    "비교",
]
for qtype in target_types:
    if qtype not in qa_by_type:
        continue
    items = qa_by_type[qtype]
    sampled = random.sample(items, min(2, len(items)))
    for item in sampled:
        # 실서비스와 동일하게 규정 원문 전문을 input에 포함 (RAG가 붙여주는 형태)
        article_full = find_article_text(item["article"])
        if article_full:
            qa_input = f"관련 규정: {item['article']}\n{article_full}\n\n질문: {item['input_q']}"
        else:
            qa_input = f"관련 규정: {item['article']}\n\n질문: {item['input_q']}"

        testset.append(
            {
                "test_id": f"BENCH-{test_id:03d}",
                "category": "regulation_qa",
                "subcategory": item["qtype"],
                "instruction": item["instruction"],
                "input": qa_input,
                "reference_output": item["output_a"],
                "metadata": {
                    "source_no": item["no"],
                    "article": item["article"],
                    "chapter": item["chapter"],
                    "question_type": item["qtype"],
                },
                "eval_criteria": [
                    "답변 정확성 (규정 내용과 일치)",
                    "설명 충실도 (예시 포함 여부)",
                    "한국어 자연스러움",
                ],
            }
        )
        test_id += 1

print(
    f"  => {len([t for t in testset if t['category'] == 'regulation_qa'])}개 생성"
)


# ============================================================
# 3. 회의록 분석 (Meeting Analysis) - 16개
# ============================================================
print("[3/4] 회의록 분석 샘플링...")

proc_dir = os.path.join(BASE_DIR, "data", "proceedings")
proc_files = sorted([f for f in os.listdir(proc_dir) if f.endswith(".json")])

for fname in proc_files:
    with open(os.path.join(proc_dir, fname), "r", encoding="utf-8") as f:
        data = json.load(f)

    sampled = random.sample(data, min(2, len(data)))
    for item in sampled:
        # output이 문자열인 경우 JSON 파싱
        output = item.get("output", "")
        if isinstance(output, str):
            try:
                output = json.loads(output)
            except json.JSONDecodeError:
                pass

        testset.append(
            {
                "test_id": f"BENCH-{test_id:03d}",
                "category": "meeting_analysis",
                "subcategory": fname.replace(".json", ""),
                "instruction": item.get(
                    "instruction",
                    "다음 회의록에서 회의 정보, 결정사항, Action Item, 참석자를 추출하세요.",
                ),
                "input": item.get("input", ""),
                "reference_output": output
                if isinstance(output, str)
                else json.dumps(output, ensure_ascii=False),
                "metadata": {"source_file": fname},
                "eval_criteria": [
                    "JSON 구조화 능력",
                    "핵심 정보 추출 정확도 (회의정보, 결정사항, action_items, 참석자)",
                    "한국어 요약 품질",
                ],
            }
        )
        test_id += 1

print(
    f"  => {len([t for t in testset if t['category'] == 'meeting_analysis'])}개 생성"
)


# ============================================================
# 4. 한국어 이해력 + 구조화 출력 테스트 - 34개
#    doc_summary 12 / risk_detection 12 / korean_understanding 10
# ============================================================
print("[4/4] 한국어 이해력 + 구조화 출력 테스트 추가...")

extra_tests = [
    # --- 문서 요약 (5개) ---
    {
        "category": "doc_summary",
        "subcategory": "보고서형",
        "instruction": "주어진 필드에 맞게 문서를 요약하세요.",
        "input": "필드: title, summary, key_points, conclusion\n\n원본 문서:\n제목: 2026년 1분기 IT 인프라 점검 보고서\n\n1. 점검 개요\n점검 기간: 2026.01.13~01.24\n대상: 전사 서버 48대, 네트워크 장비 120대\n\n2. 주요 발견사항\n- 서버 가동률 99.7% (목표 99.5% 초과 달성)\n- 디스크 사용률 80% 초과 서버 5대 발견\n- 방화벽 정책 미갱신 3건\n- SSL 인증서 만료 예정 2건 (3월 내)\n\n3. 조치 계획\n디스크 증설 및 방화벽 정책 갱신 예정",
        "reference_output": '{"title": "2026년 1분기 IT 인프라 점검 보고서", "summary": "서버 가동률 99.7%로 목표 달성. 디스크 사용률 초과 서버 5대, 방화벽 미갱신 3건 등 일부 조치 필요", "key_points": "- 서버 48대, 네트워크 장비 120대 점검\\n- 디스크 80% 초과 서버 5대\\n- 방화벽 정책 미갱신 3건\\n- SSL 인증서 만료 예정 2건", "conclusion": "전반적 안정성 양호. 디스크 증설 및 방화벽 정책 갱신, SSL 인증서 갱신 조치 필요"}',
        "eval_criteria": [
            "JSON 출력 형식 준수",
            "필드별 요약 정확도",
            "핵심 수치 포함 여부",
        ],
    },
    {
        "category": "doc_summary",
        "subcategory": "기획서형",
        "instruction": "주어진 필드에 맞게 문서를 요약하세요.",
        "input": "필드: title, background, summary, key_points\n\n원본 문서:\n제목: 사내 AI 챗봇 도입 기획서\n\n1. 배경\n현재 인사/IT/총무 부서에 월 평균 620건의 반복 문의가 발생하고 있으며, 담당자 1인당 일 평균 45분을 문의 응대에 사용하고 있습니다.\n\n2. 제안\nLLM 기반 사내 챗봇을 도입하여 규정 질문, 일정 관리, 문서 작성을 자동화합니다.\n주요 기능: 규정 판단, 문서 요약/생성, 일정 등록, 회의록 정리\n\n3. 기대 효과\n- 반복 문의 70% 자동 처리\n- 담당자 업무시간 주당 3.5시간 절감\n- 신입사원 온보딩 기간 단축",
        "reference_output": '{"title": "사내 AI 챗봇 도입 기획서", "background": "인사/IT/총무 부서에 월 620건 반복 문의 발생, 담당자 1인당 일 45분 응대에 소요", "summary": "LLM 기반 사내 챗봇으로 규정 질문, 일정 관리, 문서 작성 자동화를 제안하는 기획", "key_points": "- 규정 판단, 문서 요약/생성, 일정 등록, 회의록 정리 기능\\n- 반복 문의 70% 자동 처리 목표\\n- 주당 3.5시간 업무시간 절감 기대"}',
        "eval_criteria": [
            "JSON 출력 형식 준수",
            "배경 정보 정확 요약",
            "핵심 수치 포함 여부",
        ],
    },
    {
        "category": "doc_summary",
        "subcategory": "감사형",
        "instruction": "주어진 필드에 맞게 문서를 요약하세요.",
        "input": "필드: title, summary, key_points, action_required\n\n원본 문서:\n제목: 2025년 하반기 개인정보 처리 감사 보고서\n\n1. 감사 범위: 전사 개인정보 처리 현황 (고객DB, 직원DB, 협력사DB)\n2. 주요 결과\n- 고객 DB 접근 로그 미보관: 마케팅팀 3건\n- 퇴사자 개인정보 미파기: 인사팀 8건 (30일 초과 보관)\n- 동의 없는 제3자 제공: 영업팀 1건\n- 암호화 미적용: 협력사 DB 연동 구간 1건\n3. 전체 준수율: 87% (전년 대비 +5%p)",
        "reference_output": '{"title": "2025년 하반기 개인정보 처리 감사 보고서", "summary": "전사 개인정보 처리 감사 결과 준수율 87%. 접근로그 미보관, 퇴사자 정보 미파기 등 13건 미준수 발견", "key_points": "- 고객 DB 접근 로그 미보관 3건 (마케팅팀)\\n- 퇴사자 개인정보 미파기 8건 (인사팀)\\n- 동의 없는 제3자 제공 1건 (영업팀)\\n- 암호화 미적용 1건 (협력사 DB)", "action_required": "- 마케팅팀: 접근 로그 보관 체계 구축\\n- 인사팀: 퇴사자 정보 파기 프로세스 점검\\n- 영업팀: 제3자 제공 동의 절차 재정비\\n- IT팀: 협력사 DB 구간 암호화 적용"}',
        "eval_criteria": [
            "JSON 출력 형식 준수",
            "부서별 이슈 정확 분류",
            "조치사항 구체성",
        ],
    },
    {
        "category": "doc_summary",
        "subcategory": "공지형",
        "instruction": "주어진 필드에 맞게 문서를 요약하세요.",
        "input": "필드: title, summary, recipients, action_required\n\n원본 문서:\n[공지] 2026년 상반기 정보보호 교육 안내\n\n전 직원 대상으로 정보보호 교육을 실시합니다.\n\n1. 교육 기간: 2026.03.03 ~ 03.14 (2주간)\n2. 대상: 전 직원 (정규직, 계약직, 인턴 포함)\n3. 방법: 온라인 (사내 LMS)\n4. 소요시간: 약 2시간\n5. 필수 이수: 미이수 시 인사 불이익\n6. 문의: 보안팀 (내선 1234)",
        "reference_output": '{"title": "2026년 상반기 정보보호 교육 안내", "summary": "2026.03.03~03.14 전 직원 대상 온라인 정보보호 교육 실시. 약 2시간 소요, 미이수 시 인사 불이익", "recipients": "전 직원 (정규직, 계약직, 인턴 포함)", "action_required": "- 기간 내 LMS에서 정보보호 교육 이수 (2시간)\\n- 미이수 시 인사 불이익 있으므로 반드시 완료"}',
        "eval_criteria": [
            "JSON 출력 형식 준수",
            "일정/대상 정확도",
            "필수 조치사항 포함",
        ],
    },
    {
        "category": "doc_summary",
        "subcategory": "기술문서형",
        "instruction": "주어진 필드에 맞게 문서를 요약하세요.",
        "input": "필드: title, summary, key_points, technical_details\n\n원본 문서:\n제목: API 게이트웨이 마이그레이션 계획서\n\n1. 현황: 레거시 API 게이트웨이(Kong 2.x)의 EOL이 2026년 6월로 예정\n2. 목표: Kong 3.x + Kubernetes Ingress로 전환\n3. 영향 범위: 내부 API 47개, 외부 연동 API 12개, 일 평균 트래픽 150만 req\n4. 마이그레이션 전략\n- Blue-Green 배포 방식\n- 단계별 전환 (내부 API → 외부 API 순서)\n- 롤백 계획 포함\n5. 일정: 4월 PoC → 5월 내부 전환 → 6월 외부 전환",
        "reference_output": '{"title": "API 게이트웨이 마이그레이션 계획서", "summary": "Kong 2.x EOL 대응으로 Kong 3.x + K8s Ingress 전환 계획. Blue-Green 배포로 단계별 마이그레이션", "key_points": "- 레거시 Kong 2.x → Kong 3.x + Kubernetes Ingress\\n- 내부 47개 + 외부 12개 API 대상\\n- 4월 PoC → 5월 내부 → 6월 외부 전환", "technical_details": "- Blue-Green 배포 방식 적용\\n- 일 평균 트래픽 150만 req 처리 필요\\n- 내부 API 우선 전환 후 외부 API 순차 적용\\n- 롤백 계획 포함"}',
        "eval_criteria": [
            "JSON 출력 형식 준수",
            "기술 용어 정확성",
            "일정/수치 포함 여부",
        ],
    },
    # --- 문서 요약 추가분 (7개) ---
    {
        "category": "doc_summary",
        "subcategory": "인사공지형",
        "instruction": "주어진 필드에 맞게 문서를 요약하세요.",
        "input": "필드: title, summary, key_points, effective_date\n\n원본 문서:\n제목: 2026년 상반기 조직 개편 안내\n\n1. 개편 내용\n- AI사업부 신설 (기존 R&D팀에서 분리)\n- 마케팅팀 → 그로스팀 명칭 변경\n- 보안팀을 CISO 직속으로 격상\n\n2. 인사 이동\n- AI사업부장: 김영수 상무 (신규 선임)\n- 그로스팀장: 이지현 팀장 (유임)\n- 보안팀장: 박민호 팀장 (유임, 보고 라인 변경)\n\n3. 시행일: 2026년 3월 1일\n4. 기존 업무 인수인계는 2월 28일까지 완료",
        "reference_output": '{"title": "2026년 상반기 조직 개편 안내", "summary": "AI사업부 신설, 마케팅팀 명칭 변경(그로스팀), 보안팀 CISO 직속 격상 등 조직 개편 실시", "key_points": "- AI사업부 신설 (R&D팀에서 분리, 김영수 상무 선임)\\n- 마케팅팀 → 그로스팀 명칭 변경\\n- 보안팀 CISO 직속 격상\\n- 2월 28일까지 인수인계 완료 필요", "effective_date": "2026년 3월 1일"}',
        "eval_criteria": ["JSON 출력 형식 준수", "조직 변경사항 정확 반영", "인사 정보 정확도"],
    },
    {
        "category": "doc_summary",
        "subcategory": "회의록요약형",
        "instruction": "주어진 필드에 맞게 문서를 요약하세요.",
        "input": "필드: title, summary, decisions, action_items\n\n원본 문서:\n제목: 제5차 정보보호위원회 회의록\n일시: 2026.02.07 14:00~15:30\n참석: CISO 이강호, CPO 정미영, 개발부서장, 인사부서장, 법무팀장\n\n1. 안건: ISMS-P 인증 준비 현황\n- 현재 82개 항목 중 68개 완료 (83%)\n- 미완료 14개 중 긴급 5개: 접근권한 검토, 백업 복구 테스트, 취약점 점검, 교육 이수율, 로그 보관\n\n2. 결정사항\n- 3월 15일까지 긴급 5개 항목 완료\n- 외부 컨설팅 업체 선정 (3월 중)\n- 모의 감사 4월 실시\n\n3. 차기 회의: 2026.03.07 14:00",
        "reference_output": '{"title": "제5차 정보보호위원회 회의록", "summary": "ISMS-P 인증 준비 현황 점검. 82개 항목 중 68개 완료(83%), 긴급 미완료 5개 항목 3월 15일까지 완료 결정", "decisions": "- 긴급 5개 항목 3월 15일까지 완료\\n- 외부 컨설팅 업체 3월 중 선정\\n- 모의 감사 4월 실시", "action_items": "- 접근권한 검토, 백업 복구 테스트, 취약점 점검, 교육 이수율, 로그 보관 완료\\n- 외부 컨설팅 업체 선정 진행\\n- 차기 회의 2026.03.07 14:00"}',
        "eval_criteria": ["JSON 출력 형식 준수", "결정사항/액션아이템 구분 정확도", "수치 정보 정확도"],
    },
    {
        "category": "doc_summary",
        "subcategory": "장애보고형",
        "instruction": "주어진 필드에 맞게 문서를 요약하세요.",
        "input": "필드: title, summary, timeline, root_cause, action_taken\n\n원본 문서:\n제목: 2026.02.05 API 서버 장애 보고서\n\n1. 장애 개요\n- 발생: 2026.02.05 09:23\n- 복구: 2026.02.05 10:47 (총 84분)\n- 영향: 전사 API 서비스 응답 불가, 챗봇 서비스 중단\n- 등급: Critical\n\n2. 타임라인\n- 09:23 모니터링 알림 발생 (API 응답시간 30초 초과)\n- 09:28 보안팀 1차 확인 (서버 CPU 100%)\n- 09:45 원인 파악 (배치 작업 동시 실행으로 리소스 고갈)\n- 10:15 배치 작업 중단 및 서버 재시작\n- 10:47 서비스 정상 복구 확인\n\n3. 근본 원인: 야간 배치 작업이 09:00 시작 배치와 겹쳐 실행됨 (cron 스케줄 오류)\n4. 재발 방지: 배치 스케줄 검증 자동화, 리소스 임계치 알림 강화",
        "reference_output": '{"title": "2026.02.05 API 서버 장애 보고서", "summary": "배치 작업 충돌로 API 서버 CPU 100% → 전사 API/챗봇 84분간 중단. cron 스케줄 오류가 근본 원인", "timeline": "09:23 알림 → 09:28 1차 확인 → 09:45 원인 파악 → 10:15 배치 중단 및 재시작 → 10:47 복구 완료", "root_cause": "야간 배치 작업이 09:00 시작 배치와 겹쳐 실행 (cron 스케줄 오류)", "action_taken": "- 배치 작업 중단 및 서버 재시작으로 즉시 복구\\n- 배치 스케줄 검증 자동화 도입\\n- 리소스 임계치 알림 강화"}',
        "eval_criteria": ["JSON 출력 형식 준수", "타임라인 정확도", "근본원인/조치사항 구분"],
    },
    {
        "category": "doc_summary",
        "subcategory": "정책문서형",
        "instruction": "주어진 필드에 맞게 문서를 요약하세요.",
        "input": "필드: title, summary, scope, key_rules\n\n원본 문서:\n제목: 원격근무 보안 가이드라인 v2.0\n\n1. 목적: 원격근무 환경에서의 정보보호 강화\n2. 적용범위: 재택, 카페, 출장지 등 사외 업무 전체\n3. 필수 준수사항\n- VPN 접속 필수 (회사 지급 클라이언트만 허용)\n- 공용 Wi-Fi 사용 금지\n- 회사 지급 노트북 또는 승인된 BYOD 단말만 사용\n- 화면 잠금 5분 타이머 설정\n- 업무 화면 캡처·촬영 금지\n- 보안사고 발생 시 즉시 보안팀 보고 (내선 9119)\n4. 위반 시: 경고 → 원격근무 제한 → 징계위원회 회부",
        "reference_output": '{"title": "원격근무 보안 가이드라인 v2.0", "summary": "원격근무 시 VPN 필수, 공용 Wi-Fi 금지, 승인 단말만 사용 등 보안 수칙을 규정하는 가이드라인", "scope": "재택, 카페, 출장지 등 사외 업무 전체", "key_rules": "- VPN 접속 필수\\n- 공용 Wi-Fi 사용 금지\\n- 회사 지급 또는 승인된 BYOD 단말만 사용\\n- 화면 잠금 5분 설정\\n- 화면 캡처·촬영 금지\\n- 위반 시 경고 → 원격근무 제한 → 징계위원회 회부"}',
        "eval_criteria": ["JSON 출력 형식 준수", "규칙 항목 누락 없음", "위반 제재 정보 포함"],
    },
    {
        "category": "doc_summary",
        "subcategory": "교육결과형",
        "instruction": "주어진 필드에 맞게 문서를 요약하세요.",
        "input": "필드: title, summary, statistics, improvements\n\n원본 문서:\n제목: 2025년 하반기 정보보호 교육 결과 보고\n\n1. 교육 개요\n- 기간: 2025.10.14~10.25\n- 대상: 전 직원 52명\n- 방식: 온라인(LMS) + 피싱 시뮬레이션\n\n2. 결과\n- 이수율: 100% (52/52명)\n- 평균 점수: 84점 (전년 78점 대비 +6점)\n- 피싱 시뮬레이션 클릭률: 12% (전년 23% 대비 -11%p)\n- 최저 점수 부서: 영업팀 (평균 72점)\n\n3. 개선 필요사항\n- 영업팀 추가 교육 실시 (11월 중)\n- 모바일 피싱 시나리오 추가 필요\n- 교육 콘텐츠 최신 사례 반영 (AI 딥페이크 피싱 등)",
        "reference_output": '{"title": "2025년 하반기 정보보호 교육 결과 보고", "summary": "전 직원 52명 이수율 100%, 평균 84점(+6점), 피싱 클릭률 12%(-11%p)로 전반적 개선", "statistics": "- 이수율: 100% (52/52명)\\n- 평균 점수: 84점 (전년 78점 대비 +6점)\\n- 피싱 클릭률: 12% (전년 23% 대비 -11%p)\\n- 최저 점수 부서: 영업팀 (72점)", "improvements": "- 영업팀 추가 교육 실시\\n- 모바일 피싱 시나리오 추가\\n- AI 딥페이크 피싱 등 최신 사례 콘텐츠 반영"}',
        "eval_criteria": ["JSON 출력 형식 준수", "수치 비교 정확도", "개선사항 구체성"],
    },
    {
        "category": "doc_summary",
        "subcategory": "비용분석형",
        "instruction": "주어진 필드에 맞게 문서를 요약하세요.",
        "input": "필드: title, summary, cost_breakdown, recommendation\n\n원본 문서:\n제목: 클라우드 인프라 비용 분석 (2026년 1월)\n\n1. 총 비용: 1,840만원 (전월 대비 +15%)\n2. 항목별\n- 컴퓨팅(EC2): 920만원 (50%)\n- GPU 인스턴스(A100): 480만원 (26%) — 모델 학습용\n- 스토리지(S3+EBS): 210만원 (11%)\n- 데이터 전송: 130만원 (7%)\n- 기타(RDS, ElastiCache 등): 100만원 (6%)\n3. 비용 증가 원인: GPU 인스턴스 사용량 급증 (파인튜닝 실험 5회)\n4. 절감 방안: Spot Instance 활용, 미사용 인스턴스 자동 종료, Reserved Instance 검토",
        "reference_output": '{"title": "클라우드 인프라 비용 분석 (2026년 1월)", "summary": "총 1,840만원(전월 +15%). GPU 인스턴스(파인튜닝) 사용 급증이 주요 원인", "cost_breakdown": "- 컴퓨팅(EC2): 920만원 (50%)\\n- GPU(A100): 480만원 (26%)\\n- 스토리지: 210만원 (11%)\\n- 데이터 전송: 130만원 (7%)\\n- 기타: 100만원 (6%)", "recommendation": "- Spot Instance 활용으로 GPU 비용 절감\\n- 미사용 인스턴스 자동 종료 설정\\n- Reserved Instance 검토 (장기 사용 시)"}',
        "eval_criteria": ["JSON 출력 형식 준수", "비용 수치 정확도", "절감 방안 구체성"],
    },
    {
        "category": "doc_summary",
        "subcategory": "채용공고형",
        "instruction": "주어진 필드에 맞게 문서를 요약하세요.",
        "input": "필드: title, summary, requirements, benefits\n\n원본 문서:\n제목: [채용] AI 엔지니어 (경력 3~7년)\n\n1. 주요 업무\n- LLM 파인튜닝 및 서빙 파이프라인 구축\n- RAG 시스템 설계 및 최적화\n- 사내 AI 챗봇 모델 성능 개선\n\n2. 자격 요건\n- Python, PyTorch 실무 경력 3년 이상\n- LLM 파인튜닝(LoRA, QLoRA) 경험\n- RAG, Vector DB 활용 경험\n- 우대: vLLM, LangChain/LangGraph 경험\n\n3. 근무 조건\n- 연봉: 6,000~9,000만원 (경력에 따라 협의)\n- 재택근무 주 2일\n- 스톡옵션 부여\n- GPU 서버(A100) 개인 할당",
        "reference_output": '{"title": "AI 엔지니어 채용 (경력 3~7년)", "summary": "LLM 파인튜닝/RAG 전문 AI 엔지니어 채용. 사내 AI 챗봇 모델 성능 개선이 주요 업무", "requirements": "- Python, PyTorch 3년 이상\\n- LLM 파인튜닝(LoRA, QLoRA) 경험\\n- RAG, Vector DB 활용 경험\\n- 우대: vLLM, LangChain/LangGraph", "benefits": "- 연봉 6,000~9,000만원\\n- 재택근무 주 2일\\n- 스톡옵션 부여\\n- GPU 서버(A100) 개인 할당"}',
        "eval_criteria": ["JSON 출력 형식 준수", "자격요건/우대사항 구분", "복리후생 정확도"],
    },
    # --- 리스크 감지 (5개) ---
    {
        "category": "risk_detection",
        "subcategory": "위반_접근권한",
        "instruction": "다음 내용에서 규정 위반 여부를 감지하세요.",
        "input": "회의 내용: 인턴 박지호에게 프로덕션 DB 직접 접근 권한을 부여하기로 결정함. 개발 속도를 위해 필요하다는 팀장 의견 반영.\n\n관련 규정:\n제18조(데이터베이스 접근통제) 제3항 - 수습기간 중인 직원은 프로덕션 데이터베이스에 직접 접근할 수 없다.",
        "reference_output": '{"risk_detected": true, "risk_level": "high", "violation": "제18조(데이터베이스 접근통제) 제3항 위반", "detail": "수습기간 중인 인턴에게 프로덕션 DB 직접 접근 권한 부여 시도", "recommendation": "테스트 DB 환경을 제공하고, 수습 종료 후 접근 권한 재검토"}',
        "eval_criteria": [
            "위반 감지 정확도",
            "위반 조항 정확 인용",
            "대안 제시 적절성",
        ],
    },
    {
        "category": "risk_detection",
        "subcategory": "위반_비밀유지",
        "instruction": "다음 내용에서 규정 위반 여부를 감지하세요.",
        "input": "회의 내용: 마케팅팀 김대리가 경쟁사 이직 예정자와 진행 중인 프로젝트의 기술 스펙을 공유한 사실이 확인됨.\n\n관련 규정:\n제11조(비밀유지의무) 제1항 - 재직 중 취득한 회사의 영업비밀 및 기술정보를 외부에 유출하여서는 안 된다.",
        "reference_output": '{"risk_detected": true, "risk_level": "high", "violation": "제11조(비밀유지의무) 제1항 위반", "detail": "진행 중인 프로젝트 기술 스펙을 경쟁사 이직 예정자에게 유출", "recommendation": "즉시 해당 직원 면담 및 유출 범위 파악, 보안팀 보고 후 징계 절차 검토"}',
        "eval_criteria": [
            "위반 감지 정확도",
            "심각도 판단 적절성",
            "후속 조치 구체성",
        ],
    },
    {
        "category": "risk_detection",
        "subcategory": "정상_접근권한",
        "instruction": "다음 내용에서 규정 위반 여부를 감지하세요.",
        "input": "회의 내용: 정규직 백엔드 개발자 이수진에게 담당 프로젝트의 스테이징 DB 읽기 권한을 부여하기로 함. 부서장 승인 완료.\n\n관련 규정:\n제17조(정보시스템 접근통제) 제1항 - 업무 수행에 필요한 최소한의 접근 권한을 부여한다.\n제17조(정보시스템 접근통제) 제3항 - 접근 권한 부여 시 부서장의 승인을 받아야 한다.",
        "reference_output": '{"risk_detected": false, "risk_level": "none", "violation": "", "detail": "정규직 직원에게 담당 프로젝트 범위 내 최소 권한(스테이징 DB 읽기) 부여, 부서장 승인 완료로 절차 준수", "recommendation": ""}',
        "eval_criteria": [
            "정상 판단 정확도",
            "규정 준수 근거 설명",
            "불필요한 위험 경고 없음",
        ],
    },
    {
        "category": "risk_detection",
        "subcategory": "위반_원격근무",
        "instruction": "다음 내용에서 규정 위반 여부를 감지하세요.",
        "input": "회의 내용: 재택근무 중인 직원들이 카페에서 공용 Wi-Fi로 사내 시스템에 접속하는 사례가 보고됨. VPN 없이 직접 접속한 경우도 있음.\n\n관련 규정:\n제9조(원격근무) 제2항 - 원격근무 시 반드시 회사 승인 VPN을 통해 접속해야 한다.\n제9조(원격근무) 제4항 - 공공장소에서의 업무 수행 시 화면 보안필터를 사용해야 한다.",
        "reference_output": '{"risk_detected": true, "risk_level": "medium", "violation": "제9조(원격근무) 제2항 위반", "detail": "공용 Wi-Fi에서 VPN 미사용으로 사내 시스템 직접 접속", "recommendation": "전 직원 대상 원격근무 보안 수칙 재교육, VPN 미접속 시 사내 시스템 차단 정책 적용 검토"}',
        "eval_criteria": [
            "위반 감지 정확도",
            "관련 조항 모두 식별",
            "재발 방지 제안 포함",
        ],
    },
    {
        "category": "risk_detection",
        "subcategory": "정상_교육",
        "instruction": "다음 내용에서 규정 위반 여부를 감지하세요.",
        "input": "회의 내용: 보안팀에서 전 직원 대상 분기별 정보보호 교육을 정상 실시함. 이수율 98%, 미이수 2명은 출장 중으로 다음 주 보충 교육 예정.\n\n관련 규정:\n제29조(정보보호 교육) 제1항 - 전 직원은 분기별 1회 이상 정보보호 교육을 이수해야 한다.\n제29조(정보보호 교육) 제3항 - 부득이한 사유로 미이수 시 2주 이내 보충 교육을 이수해야 한다.",
        "reference_output": '{"risk_detected": false, "risk_level": "none", "violation": "", "detail": "분기별 교육 정상 실시(이수율 98%), 미이수자 2명은 출장 사유로 2주 이내 보충 교육 예정이므로 규정 준수", "recommendation": ""}',
        "eval_criteria": [
            "정상 판단 정확도",
            "보충 교육 규정 이해",
            "과잉 경고 없음",
        ],
    },
    # --- 리스크 감지 추가분 (7개) ---
    {
        "category": "risk_detection",
        "subcategory": "위반_개인정보",
        "instruction": "다음 내용에서 규정 위반 여부를 감지하세요.",
        "input": "회의 내용: 마케팅팀에서 고객 DB의 이름, 전화번호, 이메일 데이터를 외부 마케팅 대행사에 엑셀 파일로 전달하기로 함. 고객 동의는 받지 않은 상태.\n\n관련 규정:\n제18조(데이터베이스 접근통제) - 개인정보 및 민감정보가 저장된 테이블에 대해서는 접근 로그를 기록하고, 비정상적인 대량 조회·다운로드를 탐지·차단한다.\n제11조(비밀유지의무) - 고객 데이터 및 개인정보는 비밀유지 대상이다.",
        "reference_output": '{"risk_detected": true, "risk_level": "high", "violation": "제18조(데이터베이스 접근통제) 및 제11조(비밀유지의무) 위반, 개인정보 보호법 제17조(제3자 제공) 위반", "detail": "고객 동의 없이 개인정보(이름, 전화번호, 이메일)를 외부 대행사에 제공 시도", "recommendation": "고객 동의 절차 선행, 개인정보 제3자 제공 동의서 징구, 필요 최소한의 정보만 비식별화하여 제공 검토"}',
        "eval_criteria": ["위반 감지 정확도", "관련 법령 인식", "개인정보 보호 관점 분석"],
    },
    {
        "category": "risk_detection",
        "subcategory": "위반_소스코드",
        "instruction": "다음 내용에서 규정 위반 여부를 감지하세요.",
        "input": "회의 내용: 개발팀 신입 최대한이 개인 GitHub 공개 저장소에 사내 프로젝트 코드 일부를 포트폴리오용으로 업로드한 사실이 발견됨.\n\n관련 규정:\n제26조(소스코드 관리) - 소스코드 저장소에는 비밀번호, API 키, 인증서 등 민감 정보를 포함하여서는 아니 된다.\n제11조(비밀유지의무) - 소스코드 및 기술 설계서는 비밀유지 대상이다.",
        "reference_output": '{"risk_detected": true, "risk_level": "high", "violation": "제26조(소스코드 관리) 및 제11조(비밀유지의무) 위반", "detail": "사내 프로젝트 소스코드를 개인 공개 GitHub 저장소에 무단 업로드", "recommendation": "즉시 해당 저장소 삭제 요청, 유출 범위 파악(API 키 등 민감정보 포함 여부 확인), 보안팀 보고 후 징계 절차 검토, 전 직원 대상 소스코드 관리 교육 실시"}',
        "eval_criteria": ["위반 감지 정확도", "유출 영향 범위 인식", "즉시 조치사항 포함"],
    },
    {
        "category": "risk_detection",
        "subcategory": "정상_변경관리",
        "instruction": "다음 내용에서 규정 위반 여부를 감지하세요.",
        "input": "회의 내용: 인프라팀에서 방화벽 정책 변경을 요청하였고, 변경 요청서 제출 → 영향도 분석 → CAB 검토 → CISO 승인을 거쳐 금요일 야간에 적용 예정. 롤백 계획도 수립 완료.\n\n관련 규정:\n제21조(변경관리) - 정보시스템의 모든 변경은 사전에 변경관리 절차에 따라 신청·검토·승인·시행하여야 한다.\n제16조(네트워크 접근통제) - 네트워크 접근통제 정책의 변경은 정보보호 담당자의 검토 및 CISO의 승인을 받아야 한다.",
        "reference_output": '{"risk_detected": false, "risk_level": "none", "violation": "", "detail": "변경 요청서 → 영향도 분석 → CAB 검토 → CISO 승인의 정규 변경관리 절차를 모두 준수. 롤백 계획까지 수립하여 제21조 및 제16조 요건 충족", "recommendation": ""}',
        "eval_criteria": ["정상 판단 정확도", "변경관리 절차 이해", "과잉 경고 없음"],
    },
    {
        "category": "risk_detection",
        "subcategory": "위반_겸업",
        "instruction": "다음 내용에서 규정 위반 여부를 감지하세요.",
        "input": "회의 내용: 개발팀 박과장이 회사 승인 없이 외부 스타트업의 CTO로 활동하고 있으며, 근무시간 중 해당 스타트업 업무를 처리한 정황이 확인됨.\n\n관련 규정:\n제12조(겸업 금지) - 직원은 회사의 사전 서면 승인 없이 타 기업에 취업하거나 독립적인 영리 활동을 할 수 없다.\n제10조(복무의무) - 직원은 관련 법령, 본 규정 및 회사의 제반 규정·지침을 성실히 준수하여야 한다.",
        "reference_output": '{"risk_detected": true, "risk_level": "high", "violation": "제12조(겸업 금지) 위반", "detail": "회사 사전 서면 승인 없이 외부 스타트업 CTO로 활동, 근무시간 중 외부 업무 처리", "recommendation": "즉시 면담을 통한 사실 확인, 겸업 중단 요구, 인사위원회 회부 검토, 영업비밀 유출 여부 조사"}',
        "eval_criteria": ["위반 감지 정확도", "겸업 금지 규정 정확 적용", "후속 조치 구체성"],
    },
    {
        "category": "risk_detection",
        "subcategory": "위반_백업",
        "instruction": "다음 내용에서 규정 위반 여부를 감지하세요.",
        "input": "회의 내용: 인프라팀 점검 결과, 고객 DB 백업이 3개월째 실패 상태로 방치되어 있었음. 백업 실패 알림이 스팸 필터에 걸려 담당자가 인지하지 못한 것으로 확인됨.\n\n관련 규정:\n제23조(백업 및 복구관리) - 주요 데이터 및 시스템에 대한 백업 정책을 수립·이행한다. 반기 1회 이상 복구 테스트(DR 훈련)를 실시하고 결과를 기록한다.",
        "reference_output": '{"risk_detected": true, "risk_level": "high", "violation": "제23조(백업 및 복구관리) 위반", "detail": "고객 DB 백업이 3개월간 실패 상태로 방치. 백업 실패 알림이 스팸 필터에 걸려 미인지", "recommendation": "즉시 백업 복구 및 정상화, 알림 채널 점검(스팸 필터 예외 설정), 백업 상태 모니터링 대시보드 구축, DR 테스트 실시"}',
        "eval_criteria": ["위반 감지 정확도", "장기 방치의 심각성 인식", "재발 방지 방안 구체성"],
    },
    {
        "category": "risk_detection",
        "subcategory": "정상_오픈소스",
        "instruction": "다음 내용에서 규정 위반 여부를 감지하세요.",
        "input": "회의 내용: 개발팀 이대리가 팀장 사전 승인을 받고 회사 명의로 오픈소스 프로젝트에 버그 수정 PR을 제출함. 회사 영업비밀 미포함 확인 완료, CLA 서명은 법무팀 검토 후 진행.\n\n관련 규정:\n제27조(오픈소스 사용 및 기여 정책) - 회사 명의로 외부 오픈소스 프로젝트에 기여하는 경우: ① 팀장 사전 승인, ② 회사 영업비밀 및 독점 코드 미포함 확인, ③ CLA 서명 필요 시 법무팀 검토.",
        "reference_output": '{"risk_detected": false, "risk_level": "none", "violation": "", "detail": "팀장 사전 승인, 영업비밀 미포함 확인, CLA 법무팀 검토 진행 등 제27조의 기여 절차를 모두 준수", "recommendation": ""}',
        "eval_criteria": ["정상 판단 정확도", "오픈소스 기여 절차 이해", "과잉 경고 없음"],
    },
    {
        "category": "risk_detection",
        "subcategory": "위반_개발환경",
        "instruction": "다음 내용에서 규정 위반 여부를 감지하세요.",
        "input": "회의 내용: QA팀에서 테스트 편의를 위해 프로덕션 DB의 고객 데이터를 마스킹 없이 테스트 환경으로 복사하여 사용 중인 사실이 확인됨.\n\n관련 규정:\n제18조(데이터베이스 접근통제) - 개발·테스트 환경에서는 실 운영 데이터의 사용을 금지하며, 불가피한 경우 비식별화(마스킹, 가명처리) 후 사용한다.\n제25조(소프트웨어 개발 보안) - 개발 환경(개발·테스트·스테이징·운영)은 각각 분리하여 운영한다.",
        "reference_output": '{"risk_detected": true, "risk_level": "high", "violation": "제18조(데이터베이스 접근통제) 위반", "detail": "프로덕션 고객 데이터를 마스킹 없이 테스트 환경에 복사하여 사용", "recommendation": "즉시 테스트 환경의 실 데이터 삭제, 마스킹/가명처리된 테스트 데이터셋 생성 후 교체, 데이터 복사 절차에 비식별화 의무화"}',
        "eval_criteria": ["위반 감지 정확도", "데이터 비식별화 원칙 이해", "즉시 조치 포함"],
    },
    # --- 한국어 이해력 (7개) ---
    {
        "category": "korean_understanding",
        "subcategory": "문맥추론",
        "instruction": "다음 문장을 읽고 질문에 답하세요.",
        "input": "회사 규정에 따르면, 연차는 입사 1년 후부터 15일이 부여되며, 3년 이상 근속 시 2년마다 1일씩 가산됩니다. 김대리는 입사 7년차입니다.\n\n질문: 김대리의 연차 일수는 며칠인가요?",
        "reference_output": "김대리는 입사 7년차이므로 기본 15일에 가산 연차를 더합니다. 3년 이상 근속 시 2년마다 1일 가산이므로, 3~4년 +1일, 5~6년 +1일 = 총 2일 가산. 따라서 15 + 2 = 17일입니다.",
        "eval_criteria": ["계산 정확도", "논리적 추론 과정", "한국어 설명 자연스러움"],
    },
    {
        "category": "korean_understanding",
        "subcategory": "요약",
        "instruction": "다음 내용을 3줄 이내로 요약하세요.",
        "input": "듀듀테크놀로지는 2024년 설립된 IT 기업으로, 사내 업무 자동화 솔루션을 개발하고 있습니다. 주요 제품은 AI 기반 사내 챗봇으로, 규정 질문 답변, 문서 자동 생성, 일정 관리 등의 기능을 제공합니다. 현재 직원 수는 약 50명이며, 서울 강남구에 본사를 두고 있습니다. 최근 시리즈 A 투자를 유치하여 사업 확장을 준비 중입니다.",
        "reference_output": "듀듀테크놀로지는 2024년 설립된 IT 기업으로 AI 기반 사내 챗봇(규정 답변, 문서 생성, 일정 관리)을 개발합니다. 직원 50명 규모이며 서울 강남구에 위치합니다. 최근 시리즈 A 투자를 유치하여 사업 확장을 준비 중입니다.",
        "eval_criteria": ["핵심 정보 포함", "3줄 이내 준수", "불필요한 정보 제거"],
    },
    {
        "category": "korean_understanding",
        "subcategory": "지시따르기",
        "instruction": "아래 조건에 맞게 이메일을 작성하세요.",
        "input": "조건:\n- 수신: 개발팀 전체\n- 발신: 보안팀 박팀장\n- 내용: 다음 주 수요일(3/5) 오후 2시 보안 패치 적용으로 인해 사내 시스템 2시간 점검 예정\n- 톤: 공식적, 정중",
        "reference_output": "제목: [공지] 사내 시스템 정기 점검 안내 (3/5 수요일)\n\n개발팀 여러분께,\n\n안녕하세요. 보안팀 박팀장입니다.\n\n다음 주 수요일(3월 5일) 오후 2시부터 약 2시간 동안 보안 패치 적용을 위한 사내 시스템 정기 점검이 진행될 예정입니다.\n\n점검 시간 동안 사내 시스템 접속이 일시적으로 제한될 수 있으니, 중요 작업은 점검 전에 저장해 주시기 바랍니다.\n\n불편을 드려 죄송하며, 문의 사항이 있으시면 보안팀으로 연락 부탁드립니다.\n\n감사합니다.\n보안팀 박팀장 드림",
        "eval_criteria": [
            "조건 충족 (일시, 내용, 톤)",
            "이메일 형식 준수",
            "한국어 비즈니스 문체",
        ],
    },
    {
        "category": "korean_understanding",
        "subcategory": "분류",
        "instruction": "다음 문장들의 의도를 분류하세요. 가능한 의도: judgment, doc_search, doc_generate, meeting_generate, schedule_add, schedule_view, general",
        "input": '1. "인턴이 프로덕션 서버에 접속해도 되나요?"\n2. "이번 주 일정 알려줘"\n3. "오늘 회의 내용 정리해줘"\n4. "보안 규정 제17조 내용이 뭐야?"\n5. "이 내용으로 보고서 만들어줘"',
        "reference_output": "1. judgment - 규정상 허용 여부 판단 요청\n2. schedule_view - 일정 조회 요청\n3. meeting_generate - 회의록 정리 요청\n4. doc_search - 규정 내용 검색 요청\n5. doc_generate - 문서 생성 요청",
        "eval_criteria": [
            "5개 분류 정확도",
            "분류 근거 설명",
            "경계 케이스 처리",
        ],
    },
    {
        "category": "korean_understanding",
        "subcategory": "논리판단",
        "instruction": "다음 상황을 읽고 논리적으로 판단하세요.",
        "input": "A팀 팀장이 퇴사 예정인 B 직원의 업무 인수인계를 위해, B 직원의 모든 시스템 접근 권한을 C 직원(신입)에게 일괄 이전하려 합니다.\n\n질문: 이 조치에 어떤 문제가 있을 수 있나요?",
        "reference_output": "여러 가지 문제가 있습니다:\n\n1. 최소 권한 원칙 위반: B 직원의 모든 권한을 일괄 이전하면 C 직원(신입)에게 업무에 불필요한 권한까지 부여될 수 있습니다.\n2. 신입사원 권한 제한: 수습기간 중인 신입사원에게 고급 권한(프로덕션 DB 등)이 부여될 위험이 있습니다.\n3. 권한 검토 누락: 개별 권한을 검토하지 않고 일괄 이전하면 보안 취약점이 발생할 수 있습니다.\n\n권장 조치: 업무별로 필요한 권한만 선별 부여하고, 부서장 및 보안팀 승인을 받아야 합니다.",
        "eval_criteria": [
            "보안 관점 문제 식별",
            "다각적 분석",
            "실질적 대안 제시",
        ],
    },
    {
        "category": "korean_understanding",
        "subcategory": "존댓말변환",
        "instruction": "다음 반말 문장을 비즈니스 존댓말로 변환하세요.",
        "input": '1. "이거 내일까지 해야 돼"\n2. "그 파일 나한테 보내줘"\n3. "회의 시간 바꿔야 할 것 같아"\n4. "이번 달 실적 별로야"',
        "reference_output": '1. "해당 건은 내일까지 완료해야 합니다."\n2. "해당 파일을 저에게 전달해 주시겠습니까?"\n3. "회의 시간을 변경해야 할 것 같습니다."\n4. "이번 달 실적이 다소 저조한 상황입니다."',
        "eval_criteria": [
            "존댓말 변환 정확성",
            "비즈니스 문체 적절성",
            "원래 의미 유지",
        ],
    },
    {
        "category": "korean_understanding",
        "subcategory": "다중조건추론",
        "instruction": "다음 조건을 모두 고려하여 답하세요.",
        "input": "조건:\n- 팀장 이상은 VPN 없이 사내 시스템 접속 가능 (긴급 시)\n- 대리 이하는 반드시 VPN 필요\n- 해외 출장 시 모든 직급이 VPN + MFA 필수\n- 현재 상황: 과장 김철수가 해외 출장 중 호텔에서 VPN 없이 접속 시도\n\n질문: 김 과장의 접속이 허용되나요? 이유를 설명하세요.",
        "reference_output": "허용되지 않습니다.\n\n김 과장은 과장급이므로 일반적으로는 팀장 이상 조건에 해당하지 않아 VPN이 필요합니다. 또한 현재 해외 출장 중이므로, 직급과 관계없이 VPN + MFA가 모두 필수입니다.\n\n따라서 두 가지 조건 모두에서 VPN 없는 접속은 불허됩니다:\n1. 과장급은 팀장 이상이 아니므로 VPN 필수\n2. 해외 출장 시 모든 직급 VPN + MFA 필수",
        "eval_criteria": [
            "다중 조건 정확 적용",
            "논리적 추론 과정",
            "최종 판단 정확성",
        ],
    },
    # --- 한국어 이해력 추가분 (3개) ---
    {
        "category": "korean_understanding",
        "subcategory": "규정해석추론",
        "instruction": "다음 규정 조항을 읽고 질문에 답하세요.",
        "input": "규정 조항:\n제6조 (수습기간) - 신규 채용자에 대하여 입사일로부터 3개월간을 수습기간으로 한다.\n제17조 (정보시스템 접근통제) - 접근 권한은 분기 1회 이상 정기적으로 검토하며, 인사 변동 발생 시 즉시 권한을 변경 또는 회수한다.\n제29조 (정보보호 교육) - 신규 입사자는 입사 후 1개월 이내에 정보보호 교육을 이수하여야 하며, 이수하지 않은 경우 정보시스템 접근 권한을 제한할 수 있다.\n\n질문: 2월 1일에 입사한 신입사원이 정보보호 교육을 아직 이수하지 않았고, 현재 2월 25일입니다. 이 신입사원의 시스템 접근 권한은 어떻게 되어야 하나요?",
        "reference_output": "현재 2월 25일이므로 입사 후 약 25일이 경과한 상태입니다.\n\n1. 수습기간: 제6조에 따라 3개월(~5월 1일)까지 수습기간 중입니다.\n2. 정보보호 교육: 제29조에 따라 입사 후 1개월 이내(~3월 1일)까지 교육을 이수해야 합니다. 아직 기한이 남아 있으므로 위반은 아닙니다.\n3. 접근 권한: 현재는 교육 미이수가 규정 위반이 아니므로 즉시 권한 제한 대상은 아닙니다. 다만 3월 1일까지 미이수 시 접근 권한 제한이 가능합니다.\n\n결론: 현재는 정상이나, 3월 1일 전까지 교육 이수를 독려해야 합니다.",
        "eval_criteria": ["다중 조항 종합 해석", "날짜 기반 추론 정확성", "현실적 판단"],
    },
    {
        "category": "korean_understanding",
        "subcategory": "비교분석",
        "instruction": "두 상황을 비교하여 규정 준수 여부를 분석하세요.",
        "input": "상황 A: 개발팀 김대리가 긴급 장애 대응을 위해 운영 서버의 코드를 직접 수정하고, 사후에 변경관리 절차를 24시간 내에 완료했다.\n\n상황 B: 개발팀 박과장이 새로운 기능 배포를 위해 변경관리 절차 없이 운영 서버에 직접 코드를 배포했다. 사후 절차도 진행하지 않았다.\n\n관련 규정: 제25조(소프트웨어 개발 보안) - 운영 환경에 대한 직접 코드 수정을 금지한다. 제21조(변경관리) - 긴급 변경의 경우 사후 승인을 허용하되, 변경 시행 후 24시간 이내에 변경관리 절차를 완료하여야 한다.",
        "reference_output": "상황 A: 규정 일부 위반이나 참작 가능\n- 제25조의 운영 환경 직접 코드 수정 금지 원칙에는 위반됩니다.\n- 다만 긴급 장애 대응 상황이었고, 제21조에 따라 사후 24시간 내 변경관리 절차를 완료했으므로 긴급 변경 예외 조항을 적용할 수 있습니다.\n\n상황 B: 명백한 규정 위반\n- 제25조의 운영 환경 직접 코드 수정 금지 위반입니다.\n- 긴급 상황이 아닌 신규 기능 배포이므로 긴급 변경 예외를 적용할 수 없습니다.\n- 사후 변경관리 절차도 미진행으로 제21조도 위반입니다.\n\n결론: 상황 A는 긴급 예외 적용 가능, 상황 B는 이중 위반으로 징계 대상입니다.",
        "eval_criteria": ["두 상황 비교 정확도", "예외 조항 적용 능력", "규정 근거 명시"],
    },
    {
        "category": "korean_understanding",
        "subcategory": "절차나열",
        "instruction": "주어진 규정을 바탕으로 올바른 절차를 순서대로 나열하세요.",
        "input": "다음 조항들을 참고하여, 신입사원이 입사 후 사내 시스템에 정상적으로 접근하기까지 거쳐야 하는 절차를 순서대로 정리하세요.\n\n제4조(채용) - 채용 시 비밀유지서약서를 제출하여야 한다.\n제5조(근로계약) - 정보보호 의무 및 비밀유지 조항이 포함된 근로계약을 체결한다.\n제6조(수습기간) - 입사일로부터 3개월간 수습기간이다.\n제17조(정보시스템 접근통제) - 업무상 필요한 최소한의 접근 권한을 부여하며, 부서장 승인이 필요하다.\n제29조(정보보호 교육) - 입사 후 1개월 이내에 정보보호 교육을 이수해야 한다.",
        "reference_output": "신입사원 사내 시스템 접근까지의 절차:\n\n1. 비밀유지서약서 제출 (제4조) — 채용 확정 시\n2. 근로계약 체결 (제5조) — 정보보호 의무 조항 포함\n3. 수습기간 시작 (제6조) — 입사일부터 3개월\n4. 부서장 승인을 받아 최소 권한 접근 권한 부여 (제17조)\n5. 입사 후 1개월 이내 정보보호 교육 이수 (제29조)\n\n참고: 정보보호 교육 미이수 시 시스템 접근 권한이 제한될 수 있으며(제29조), 수습기간 중 정보보호 규정 준수 여부가 평가 기준에 포함됩니다(제6조).",
        "eval_criteria": ["절차 순서 정확성", "근거 조항 매핑", "보충 설명 적절성"],
    },
]

for item in extra_tests:
    testset.append(
        {
            "test_id": f"BENCH-{test_id:03d}",
            "category": item["category"],
            "subcategory": item["subcategory"],
            "instruction": item["instruction"],
            "input": item["input"],
            "reference_output": item["reference_output"],
            "metadata": {"source": "manual"},
            "eval_criteria": item["eval_criteria"],
        }
    )
    test_id += 1

extra_categories = {"doc_summary", "risk_detection", "korean_understanding"}
for cat in extra_categories:
    count = len([t for t in testset if t["category"] == cat])
    print(f"  => {cat}: {count}개 생성")


# ============================================================
# 저장
# ============================================================
os.makedirs(os.path.dirname(OUTPUT_PATH), exist_ok=True)
with open(OUTPUT_PATH, "w", encoding="utf-8") as f:
    for item in testset:
        f.write(json.dumps(item, ensure_ascii=False) + "\n")

print(f"\n{'='*50}")
print(f"총 {len(testset)}개 테스트 항목 생성 완료")
print(f"저장: {OUTPUT_PATH}")

# 카테고리별 요약
print(f"\n카테고리별 분포:")
cats = {}
for item in testset:
    cats[item["category"]] = cats.get(item["category"], 0) + 1
for k, v in sorted(cats.items()):
    print(f"  {k}: {v}개")
